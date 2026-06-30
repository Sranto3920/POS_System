"""
ExportService — shop-scoped Excel exports using openpyxl.
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import text

from extensions import db


class ExportService:
    @staticmethod
    def _save_workbook(wb, prefix):
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return buffer, f"{prefix}_{timestamp}.xlsx"

    @staticmethod
    def _write_sheet(ws, headers, rows):
        header_font = Font(bold=True)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    @staticmethod
    def export_products(shop_id):
        rows = db.session.execute(
            text(
                """
                SELECT
                  p.product_name,
                  c.category_name,
                  sup.supplier_name,
                  p.cost_price,
                  p.default_selling_price,
                  p.stock_quantity,
                  p.expiry_date
                FROM Products p
                LEFT JOIN Categories c
                  ON p.category_id = c.category_id AND c.shop_id = :shop_id
                LEFT JOIN Suppliers sup
                  ON p.supplier_id = sup.supplier_id AND sup.shop_id = :shop_id
                WHERE p.shop_id = :shop_id
                ORDER BY p.product_name ASC
                """
            ),
            {"shop_id": shop_id},
        ).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ExportService._write_sheet(
            ws,
            [
                "Product",
                "Category",
                "Supplier",
                "Cost Price",
                "Selling Price",
                "Stock",
                "Expiry Date",
            ],
            rows,
        )
        return ExportService._save_workbook(wb, "products")

    @staticmethod
    def export_customers(shop_id):
        rows = db.session.execute(
            text(
                """
                SELECT customer_name, phone, email, address
                FROM Customers
                WHERE shop_id = :shop_id
                ORDER BY customer_name ASC
                """
            ),
            {"shop_id": shop_id},
        ).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        ExportService._write_sheet(
            ws, ["Name", "Phone", "Email", "Address"], rows
        )
        return ExportService._save_workbook(wb, "customers")

    @staticmethod
    def export_suppliers(shop_id):
        rows = db.session.execute(
            text(
                """
                SELECT supplier_name, phone, email, address
                FROM Suppliers
                WHERE shop_id = :shop_id
                ORDER BY supplier_name ASC
                """
            ),
            {"shop_id": shop_id},
        ).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Suppliers"
        ExportService._write_sheet(
            ws, ["Name", "Phone", "Email", "Address"], rows
        )
        return ExportService._save_workbook(wb, "suppliers")

    @staticmethod
    def export_sales(shop_id):
        rows = db.session.execute(
            text(
                """
                SELECT
                  s.sale_id,
                  s.sale_date,
                  COALESCE(c.customer_name, 'Walk-in') AS customer_name,
                  u.name AS cashier_name,
                  s.total_amount
                FROM Sales s
                LEFT JOIN Customers c
                  ON s.customer_id = c.customer_id AND c.shop_id = :shop_id
                LEFT JOIN Users u
                  ON s.user_id = u.user_id AND u.shop_id = :shop_id
                WHERE s.shop_id = :shop_id
                ORDER BY s.sale_date DESC, s.sale_id DESC
                """
            ),
            {"shop_id": shop_id},
        ).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"
        ExportService._write_sheet(
            ws,
            ["Sale ID", "Date", "Customer", "Cashier", "Total Amount"],
            rows,
        )
        return ExportService._save_workbook(wb, "sales")

    @staticmethod
    def export_purchases(shop_id):
        rows = db.session.execute(
            text(
                """
                SELECT
                  p.purchase_id,
                  p.purchase_date,
                  s.supplier_name,
                  u.name AS recorded_by,
                  p.total_amount
                FROM Purchases p
                LEFT JOIN Suppliers s
                  ON p.supplier_id = s.supplier_id AND s.shop_id = :shop_id
                LEFT JOIN Users u
                  ON p.user_id = u.user_id AND u.shop_id = :shop_id
                WHERE p.shop_id = :shop_id
                ORDER BY p.purchase_date DESC, p.purchase_id DESC
                """
            ),
            {"shop_id": shop_id},
        ).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Purchases"
        ExportService._write_sheet(
            ws,
            ["Purchase ID", "Date", "Supplier", "Recorded By", "Total Amount"],
            rows,
        )
        return ExportService._save_workbook(wb, "purchases")

    @staticmethod
    def export_payments(shop_id):
        rows = db.session.execute(
            text(
                """
                SELECT
                  pay.payment_id,
                  pay.payment_date,
                  pay.sale_id,
                  pay.payment_method,
                  pay.paid_amount
                FROM Payments pay
                INNER JOIN Sales s ON pay.sale_id = s.sale_id
                WHERE s.shop_id = :shop_id
                ORDER BY pay.payment_date DESC, pay.payment_id DESC
                """
            ),
            {"shop_id": shop_id},
        ).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Payments"
        ExportService._write_sheet(
            ws,
            ["Payment ID", "Date", "Sale ID", "Method", "Amount"],
            rows,
        )
        return ExportService._save_workbook(wb, "payments")

    @staticmethod
    def export_ledger(shop_id):
        rows = db.session.execute(
            text(
                """
                SELECT
                  l.ledger_id,
                  l.transaction_date,
                  c.customer_name,
                  l.sale_id,
                  l.debit,
                  l.credit,
                  l.balance,
                  l.remarks
                FROM Customer_Ledger l
                LEFT JOIN Customers c
                  ON l.customer_id = c.customer_id AND c.shop_id = :shop_id
                WHERE l.shop_id = :shop_id
                ORDER BY l.transaction_date DESC, l.ledger_id DESC
                """
            ),
            {"shop_id": shop_id},
        ).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Ledger"
        ExportService._write_sheet(
            ws,
            [
                "Ledger ID",
                "Date",
                "Customer",
                "Sale ID",
                "Debit",
                "Credit",
                "Balance",
                "Remarks",
            ],
            rows,
        )
        return ExportService._save_workbook(wb, "ledger")
